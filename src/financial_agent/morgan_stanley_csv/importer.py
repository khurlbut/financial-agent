from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from . import db


@dataclass(frozen=True)
class ImportedSnapshot:
    snapshot_id: int
    as_of: datetime
    rows_imported: int


def import_positions_csv(
    *,
    db_path: Path,
    csv_path: Path,
    as_of: datetime | None = None,
    snapshot_id: int | None = None,
    container_id: str | None = None,
    account_name_override: str | None = None,
    force_account_name_override: bool = False,
) -> ImportedSnapshot:
    """Import a Morgan Stanley positions export (.csv or .xlsx) into SQLite.

    Morgan Stanley export formats vary across portals (Wealth Mgmt, StockPlan, etc.).
    This importer is intentionally best-effort:
    - scans for a header row that contains a "Symbol" column
    - maps common column aliases (Ticker, Shares, Market Value, etc.)
    - stores both normalized fields and raw row JSON
    """

    if as_of is None:
        as_of = datetime.now(timezone.utc)

    conn = db.connect(db_path)
    try:
        if snapshot_id is None:
            cid = (container_id or "").strip() or "morgan_stanley"
            snapshot_id = db.insert_snapshot(conn, container_id=cid, as_of=as_of, csv_path=csv_path)

        rows_imported = 0

        suffix = csv_path.suffix.lower()
        if suffix == ".csv":
            rows_iter = _iter_rows_from_csv(csv_path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows_iter = _iter_rows_from_xlsx(csv_path)
        else:
            raise ValueError(f"Unsupported Morgan Stanley export type: {csv_path.name}")

        for row in rows_iter:
            if _is_empty_row(row) or _is_summary_row(row):
                continue

            norm = _normalize_row(row)
            if account_name_override:
                existing_account = norm.get("account_name")
                if force_account_name_override or existing_account is None or not str(existing_account).strip():
                    norm["account_name"] = account_name_override
            if norm.get("symbol") is None:
                continue

            conn.execute(
                """
                INSERT INTO ms_csv_positions (
                  snapshot_id, account_name, symbol, description, quantity, price, market_value, currency, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    norm.get("account_name"),
                    norm.get("symbol"),
                    norm.get("description"),
                    norm.get("quantity"),
                    norm.get("price"),
                    norm.get("market_value"),
                    norm.get("currency"),
                    json.dumps(row),
                ),
            )
            rows_imported += 1

        conn.commit()
        return ImportedSnapshot(snapshot_id=int(snapshot_id), as_of=as_of, rows_imported=rows_imported)
    finally:
        conn.close()


def _iter_rows_from_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        _guard_against_schwab_csv(f, path)
        f.seek(0)
        raw_reader = csv.reader(f)
        fieldnames: list[str] | None = None

        for row in raw_reader:
            if not row or all(not str(c).strip() for c in row):
                continue

            normalized = [_normalize_key(str(c)) for c in row]
            if "symbol" in normalized or (row and str(row[0]).strip().lower() == "symbol"):
                fieldnames = [str(c).strip() for c in row]
                while fieldnames and fieldnames[-1] == "":
                    fieldnames.pop()
                break

        if not fieldnames:
            raise ValueError("Could not find Morgan Stanley header row (expected a row containing a 'Symbol' column)")

        reader = csv.DictReader(f, fieldnames=fieldnames)
        for row in reader:
            yield row


def _iter_rows_from_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required to import .xlsx files. Install it via requirements.txt") from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active

        header: list[str] | None = None
        header_idx: int | None = None

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row is None:
                continue
            cells = ["" if c is None else str(c).strip() for c in row]
            if not any(cells):
                continue
            normalized = [_normalize_key(c) for c in cells if c]
            if "symbol" in normalized:
                header = cells
                header_idx = i
                break

        if not header or header_idx is None:
            raise ValueError("Could not find Morgan Stanley header row in .xlsx (expected a row with a 'Symbol' column)")

        keys = header
        for row in ws.iter_rows(values_only=True, min_row=header_idx + 1):
            cells = [None if c is None else str(c).strip() for c in row]
            if not any((c or "").strip() for c in cells):
                continue

            d: dict[str, Any] = {}
            for k, v in zip(keys, cells):
                if k is None:
                    continue
                k2 = str(k).strip()
                if not k2:
                    continue
                d[k2] = v
            yield d
    finally:
        try:
            wb.close()
        except Exception:
            pass


_SCHWAB_PREAMBLE_RE = re.compile(r"^\s*\"?positions\s+for\s+account\s+.+\s+as\s+of\s+", re.I)


def _guard_against_schwab_csv(file_obj, csv_path: Path) -> None:
    """Best-effort guard to avoid importing Schwab exports as Morgan Stanley.

    Schwab CSVs are common in this repo and share a similar shape (Symbol, Qty, Price, Market Value)
    so without a guard it's easy to accidentally point the Morgan importer at a Schwab file.
    """

    try:
        # Peek at the first handful of non-empty lines.
        lines: list[str] = []
        for _ in range(30):
            line = file_obj.readline()
            if not line:
                break
            if str(line).strip():
                lines.append(str(line).strip())
            if len(lines) >= 5:
                break

        if not lines:
            return

        first = lines[0]
        if not _SCHWAB_PREAMBLE_RE.search(first):
            return

        # Schwab header line often includes "Qty (Quantity)" and "Mkt Val".
        joined = "\n".join(lines).lower()
        if "qty (quantity)" in joined and "mkt val" in joined:
            raise ValueError(
                f"{csv_path.name} looks like a Schwab Positions CSV. Use `python -m financial_agent.schwab_refresh` for this file."
            )
    except Exception as exc:
        # If it's our explicit ValueError, propagate. Otherwise, don't block import.
        if isinstance(exc, ValueError):
            raise
        return


def _normalize_key(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def _get(row: dict[Any, Any], *keys: str) -> str | None:
    by_norm: dict[str, Any] = {}
    for k, v in row.items():
        if k is None:
            continue
        by_norm[_normalize_key(str(k))] = v

    for k in keys:
        v = by_norm.get(_normalize_key(k))
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return None


def _is_empty_row(row: dict[Any, Any]) -> bool:
    values = []
    for k, v in row.items():
        if k is None:
            continue
        values.append(str(v).strip() if v is not None else "")
    return not any(values)


def _is_summary_row(row: dict[Any, Any]) -> bool:
    account = _get(row, "Account", "Account Name", "Account Number")
    if account and account.strip().lower() in {"total", "grand total"}:
        return True

    symbol = _get(row, "Symbol", "Ticker")
    if symbol and symbol.strip().lower().startswith("account total"):
        return True
    return False


_MONEY_RE = re.compile(r"[^0-9\-\.]" )


def _parse_money(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = _MONEY_RE.sub("", value)
    if cleaned in {"", "-", "."}:
        return None
    try:
        d = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    return str(d)


def _parse_qty(value: str | None) -> str | None:
    return _parse_money(value)


def _normalize_row(row: dict[Any, Any]) -> dict[str, str | None]:
    account = _get(row, "Account", "Account Name", "Account Number")
    symbol = _get(row, "Symbol", "Ticker", "Security Symbol")
    description = _get(row, "Description", "Security Description", "Name")

    quantity = _parse_qty(_get(row, "Quantity", "Shares", "Qty", "Units"))
    price = _parse_money(_get(row, "Price", "Last Price", "Market Price", "Current Price"))
    market_value = _parse_money(_get(row, "Market Value", "Value", "Current Value", "MarketValue"))

    currency = _get(row, "Currency", "Base Currency")
    if currency is not None:
        currency = currency.strip().upper()

    security_type = _get(row, "Security Type", "Type", "Asset Class")

    # Cash handling.
    if not symbol and description and "cash" in description.lower():
        symbol = "USD"

    if symbol and "cash" in symbol.lower():
        symbol = "USD"
        description = "Cash"
        if quantity is None:
            quantity = market_value
        if price is None:
            price = "1"

    if security_type and "cash" in security_type.lower() and (symbol or "").upper() in {"USD", "USDC"}:
        if quantity is None:
            quantity = market_value
        if price is None:
            price = "1"

    return {
        "account_name": account,
        "symbol": symbol.strip().upper() if symbol else None,
        "description": description,
        "quantity": quantity,
        "price": price,
        "market_value": market_value,
        "currency": currency or "USD",
    }
